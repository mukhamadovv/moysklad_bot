import io
from datetime import datetime, timezone as dt_timezone
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from bot.models import Customer, Transaction
from bot.moysklad_api import get_demand_positions, get_product


# ── Styles ────────────────────────────────────────────────────────────────────

_thin = Side(style="thin", color="BFBFBF")
_thick = Side(style="medium", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_outer_border = Border(left=_thick, right=_thick, top=_thick, bottom=_thick)

_header_font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_header_fill      = PatternFill(start_color="2E5FA3", end_color="2E5FA3", fill_type="solid")
_header_align     = Alignment(horizontal="center", vertical="center", wrap_text=True)

_title_font       = Font(name="Calibri", bold=True, size=15, color="1F3864")
_subtitle_font    = Font(name="Calibri", size=10, italic=True, color="808080")

_data_font        = Font(name="Calibri", size=10)
_money_fmt        = '#,##0.00'

_neg_font         = Font(name="Calibri", size=10, color="C00000")
_pos_font         = Font(name="Calibri", size=10, color="375623")
_warn_font        = Font(name="Calibri", size=10, color="C55A11")   # orange for pending

_center           = Alignment(horizontal="center", vertical="center")
_left             = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_right            = Alignment(horizontal="right",  vertical="center")

_stripe_fill      = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
_summary_font     = Font(name="Calibri", bold=True, size=11)
_summary_fill     = PatternFill(start_color="D6E4F7", end_color="D6E4F7", fill_type="solid")
_total_fill       = PatternFill(start_color="2E5FA3", end_color="2E5FA3", fill_type="solid")
_total_font       = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

_cust_header_fill = PatternFill(start_color="E9F0FB", end_color="E9F0FB", fill_type="solid")
_cust_header_font = Font(name="Calibri", bold=True, size=11, color="1F3864")

_balance_neg_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
_balance_pos_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def _apply_border(ws, row, col_count, border=None):
    b = border or _border
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = b


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _border


def _money(ws, row, col, value, font=None, fill=None, fmt=_money_fmt):
    cell = ws.cell(row=row, column=col, value=float(value) if value else "")
    cell.alignment = _right
    cell.border = _border
    cell.font = font or _data_font
    if fill:
        cell.fill = fill
    if isinstance(cell.value, float):
        cell.number_format = fmt
    return cell


def _text(ws, row, col, value, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align or _left
    cell.border = _border
    cell.font = font or _data_font
    if fill:
        cell.fill = fill
    return cell


def _fmt(value):
    """Format number Russian-style for display in text."""
    v = Decimal(str(value))
    sign = "−" if v < 0 else ""
    v = abs(v)
    integer_part = int(v)
    decimal_part = f"{v - integer_part:.2f}"[2:]
    s = f"{integer_part:,}".replace(",", " ")
    return f"{sign}{s},{decimal_part}"


# Cache: moysklad_entity_id → product summary string (avoids duplicate API calls)
_product_cache: dict[str, str] = {}


def _get_product_summary(tx: "Transaction") -> str:
    """Return a product name string for a sale transaction.

    Priority:
    1. Already encoded in description after ': '  (new records)
    2. Cached from a previous call for the same moysklad_entity_id
    3. Live fetch from МойСклад positions API  (old/legacy records)
    """
    desc = tx.description or ""

    # New format: "Отгрузка ОТ-ХХ: Cola ×2, Fanta ×1"  or  "Продажа: Cola ×2"
    if ": " in desc:
        return desc.split(": ", 1)[1]

    # Try cache / live API fetch
    ms_id = tx.moysklad_entity_id
    if not ms_id:
        return tx.document_number or ""

    if ms_id in _product_cache:
        return _product_cache[ms_id]

    try:
        entity_type = {
            "sale": "demand",
        }.get(tx.type, "demand")
        positions = get_demand_positions(ms_id, entity_type)
        names = []
        for pos in positions:
            assortment = pos.get("assortment", {})
            p_name = assortment.get("name", "")
            if not p_name:
                p_href = assortment.get("meta", {}).get("href", "")
                if p_href:
                    prod = get_product(p_href)
                    if prod:
                        p_name = prod.get("name", "")
            qty = float(pos.get("quantity", 0))
            if p_name:
                qty_str = f"×{int(qty)}" if qty == int(qty) else f"×{qty}"
                names.append(f"{p_name} {qty_str}")
        summary = ", ".join(names) if names else (tx.document_number or "")
    except Exception:
        summary = tx.document_number or ""

    _product_cache[ms_id] = summary
    return summary


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Purchase History Excel ────────────────────────────────────────────────────

def generate_purchase_history(customer, transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = "История операций"
    ws.sheet_properties.tabColor = "2E5FA3"

    # ── Title block ──
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"📋 История операций — {customer.full_name}")
    title_cell.font = _title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    today = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.cell(row=2, column=1,
            value=f"Сформировано: {today}  |  Телефон: {customer.phone}").font = _subtitle_font
    ws.row_dimensions[2].height = 18

    # ── Balance summary block ──
    balance = customer.debt_balance
    bonus   = customer.bonus_balance

    ws.merge_cells("A3:D3")
    bal_fill = _balance_neg_fill if balance < 0 else (_balance_pos_fill if balance > 0 else _summary_fill)
    bal_cell = ws.cell(row=3, column=1, value=f"💰 Баланс: {_fmt(balance)}")
    bal_cell.font = Font(name="Calibri", bold=True, size=11)
    bal_cell.fill = bal_fill
    bal_cell.border = _border
    for c in range(2, 5):
        ws.cell(row=3, column=c).fill = bal_fill
        ws.cell(row=3, column=c).border = _border

    ws.merge_cells("E3:H3")
    bon_cell = ws.cell(row=3, column=5, value=f"💎 Бонусы: {_fmt(bonus)}")
    bon_cell.font = Font(name="Calibri", bold=True, size=11)
    bon_cell.fill = _summary_fill
    bon_cell.border = _border
    for c in range(6, 9):
        ws.cell(row=3, column=c).fill = _summary_fill
        ws.cell(row=3, column=c).border = _border
    ws.row_dimensions[3].height = 26

    # ── Data table ──
    header_row = 5
    headers = ["№", "Дата", "Тип операции", "Документ",
               "Приход (+)", "Расход (−)", "Бонус", "Описание"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 28

    total_income  = Decimal("0")
    total_expense = Decimal("0")
    total_bonus   = Decimal("0")
    income_types  = {"payment_in", "cash_in", "return"}
    expense_types = {"sale", "payment_out", "cash_out"}

    for idx, t in enumerate(transactions, 1):
        row      = header_row + idx
        stripe   = idx % 2 == 0
        fill     = _stripe_fill if stripe else None

        is_inc = t.type in income_types
        is_exp = t.type in expense_types
        if is_inc:  total_income  += t.amount
        if is_exp:  total_expense += t.amount
        total_bonus += t.bonus_amount

        _text(ws, row, 1, idx,   align=_center, fill=fill)
        _text(ws, row, 2, t.document_date.strftime("%d.%m.%Y") if t.document_date else "",
              align=_center, fill=fill)
        _text(ws, row, 3, t.get_type_display(), fill=fill)
        _text(ws, row, 4, t.document_number, align=_center, fill=fill)
        _money(ws, row, 5, t.amount if is_inc else None,
               font=_pos_font if is_inc else _data_font, fill=fill)
        _money(ws, row, 6, t.amount if is_exp else None,
               font=_neg_font if is_exp else _data_font, fill=fill)
        _money(ws, row, 7, t.bonus_amount if t.bonus_amount else None,
               font=(_pos_font if t.bonus_amount > 0 else _neg_font) if t.bonus_amount else _data_font,
               fill=fill)
        _text(ws, row, 8, t.description, fill=fill)

    # ── Totals row ──
    n  = len(transactions)
    tr = header_row + n + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
    lbl = ws.cell(row=tr, column=1, value="ИТОГО")
    lbl.font = _summary_font; lbl.fill = _summary_fill
    lbl.alignment = _right;   lbl.border = _border
    for c in range(2, 5):
        ws.cell(row=tr, column=c).fill = _summary_fill
        ws.cell(row=tr, column=c).border = _border
    _money(ws, tr, 5, total_income,
           font=Font(name="Calibri", bold=True, size=11, color="375623"), fill=_summary_fill)
    _money(ws, tr, 6, total_expense,
           font=Font(name="Calibri", bold=True, size=11, color="C00000"), fill=_summary_fill)
    _money(ws, tr, 7, total_bonus,  font=_summary_font, fill=_summary_fill)
    ws.cell(row=tr, column=8).fill   = _summary_fill
    ws.cell(row=tr, column=8).border = _border

    _set_col_widths(ws, {"A": 5, "B": 13, "C": 20, "D": 17,
                          "E": 16, "F": 16, "G": 14, "H": 46})
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Admin Client Report ───────────────────────────────────────────────────────

def generate_admin_report(date_from: datetime, date_to: datetime) -> bytes:
    """
    Two-sheet Excel report:
      Sheet 1 – Summary: one row per customer with period totals
      Sheet 2 – Details: every sale transaction grouped by customer
    """
    # Make date_from / date_to timezone-aware (UTC)
    from django.utils import timezone as dj_tz
    if date_from.tzinfo is None:
        date_from = date_from.replace(tzinfo=dt_timezone.utc)
    if date_to.tzinfo is None:
        # include full last day
        date_to = date_to.replace(hour=23, minute=59, second=59, tzinfo=dt_timezone.utc)

    period_label = (
        f"{date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    )
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Fetch sale transactions in the period (across all customers)
    sale_txs = (
        Transaction.objects
        .filter(type="sale", document_date__range=(date_from, date_to))
        .select_related("customer")
        .order_by("customer__full_name", "document_date")
    )

    # Group by customer
    from collections import defaultdict
    by_customer: dict[int, list] = defaultdict(list)
    customer_obj: dict[int, Customer] = {}
    for tx in sale_txs:
        cid = tx.customer_id
        by_customer[cid].append(tx)
        customer_obj[cid] = tx.customer

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ═══════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.sheet_properties.tabColor = "2E5FA3"

    # Title
    SCOLS = 9
    ws1.merge_cells(f"A1:{get_column_letter(SCOLS)}1")
    ws1.cell(row=1, column=1, value=f"📊 Отчет по клиентам — {period_label}").font = _title_font
    ws1.row_dimensions[1].height = 34

    ws1.merge_cells(f"A2:{get_column_letter(SCOLS)}2")
    ws1.cell(row=2, column=1,
             value=f"Сформировано: {generated_at}  |  Клиентов в периоде: {len(by_customer)}").font = _subtitle_font
    ws1.row_dimensions[2].height = 16

    # Header row
    hr = 4
    s_headers = [
        "№", "Клиент", "Телефон", "Регион",
        "Покупок", "Сумма покупок",
        "Бонус начислен", "Бонус ожидает", "Текущий долг",
    ]
    for col, h in enumerate(s_headers, 1):
        ws1.cell(row=hr, column=col, value=h)
    _style_header(ws1, hr, SCOLS)
    ws1.row_dimensions[hr].height = 30

    grand_count  = 0
    grand_amount = Decimal("0")
    grand_earned = Decimal("0")
    grand_pend   = Decimal("0")

    for row_idx, (cid, txs) in enumerate(by_customer.items(), 1):
        cust   = customer_obj[cid]
        stripe = row_idx % 2 == 0
        fill   = _stripe_fill if stripe else None
        row    = hr + row_idx

        count  = len(txs)
        amount = sum(tx.amount        for tx in txs)
        earned = sum(tx.bonus_amount  for tx in txs)
        pend   = sum(tx.pending_bonus for tx in txs)
        debt   = cust.debt_balance

        grand_count  += count
        grand_amount += amount
        grand_earned += earned
        grand_pend   += pend

        _text(ws1, row, 1, row_idx,             align=_center, fill=fill)
        _text(ws1, row, 2, cust.full_name or "—", fill=fill,
              font=Font(name="Calibri", size=10, bold=True))
        _text(ws1, row, 3, cust.phone,          align=_center, fill=fill)
        _text(ws1, row, 4, cust.region or "—",  align=_center, fill=fill)
        _text(ws1, row, 5, count,               align=_center, fill=fill)
        _money(ws1, row, 6, amount, fill=fill)
        _money(ws1, row, 7, earned,
               font=_pos_font if earned > 0 else _data_font, fill=fill)
        _money(ws1, row, 8, pend,
               font=_warn_font if pend > 0 else _data_font, fill=fill)

        d_fill = _balance_neg_fill if debt < 0 else (_balance_pos_fill if debt > 0 else fill)
        _money(ws1, row, 9, debt,
               font=(_neg_font if debt < 0 else (_pos_font if debt > 0 else _data_font)),
               fill=d_fill)

    # Grand total
    tr = hr + len(by_customer) + 1
    ws1.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
    lbl = ws1.cell(row=tr, column=1, value="ИТОГО ЗА ПЕРИОД")
    lbl.font = _total_font; lbl.fill = _total_fill
    lbl.alignment = _right; lbl.border = _border
    for c in range(2, 5):
        ws1.cell(row=tr, column=c).font  = _total_font
        ws1.cell(row=tr, column=c).fill  = _total_fill
        ws1.cell(row=tr, column=c).border = _border
    _text(ws1, tr, 5, grand_count, align=_center,
          font=_total_font, fill=_total_fill)
    _money(ws1, tr, 6, grand_amount, font=_total_font, fill=_total_fill)
    _money(ws1, tr, 7, grand_earned, font=_total_font, fill=_total_fill)
    _money(ws1, tr, 8, grand_pend,   font=_total_font, fill=_total_fill)
    ws1.cell(row=tr, column=9).fill   = _total_fill
    ws1.cell(row=tr, column=9).border = _border

    _set_col_widths(ws1, {
        "A": 5, "B": 28, "C": 17, "D": 14,
        "E": 10, "F": 18, "G": 18, "H": 18, "I": 18,
    })
    ws1.freeze_panes = f"A{hr + 1}"

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 2 — DETAILED SALES
    # ═══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Детали продаж")
    ws2.sheet_properties.tabColor = "375623"

    DCOLS = 8
    ws2.merge_cells(f"A1:{get_column_letter(DCOLS)}1")
    ws2.cell(row=1, column=1,
             value=f"📋 Детали продаж по клиентам — {period_label}").font = _title_font
    ws2.row_dimensions[1].height = 34

    ws2.merge_cells(f"A2:{get_column_letter(DCOLS)}2")
    ws2.cell(row=2, column=1,
             value=f"Сформировано: {generated_at}").font = _subtitle_font
    ws2.row_dimensions[2].height = 16

    d_headers = [
        "№", "Дата", "Документ",
        "Товары / Описание", "Сумма",
        "Бонус начислен", "Бонус ожидает", "Долг изменение",
    ]
    dhr = 4
    for col, h in enumerate(d_headers, 1):
        ws2.cell(row=dhr, column=col, value=h)
    _style_header(ws2, dhr, DCOLS)
    ws2.row_dimensions[dhr].height = 30
    ws2.freeze_panes = f"A{dhr + 1}"

    current_row = dhr + 1
    global_idx  = 0

    for cid, txs in by_customer.items():
        cust = customer_obj[cid]

        # ── Customer section header ──────────────────────────────────
        ws2.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=DCOLS
        )
        ch = ws2.cell(
            row=current_row, column=1,
            value=f"👤  {cust.full_name or '—'}   •   {cust.phone}   •   {cust.region or '—'}"
        )
        ch.font  = _cust_header_font
        ch.fill  = _cust_header_fill
        ch.alignment = _left
        ch.border = _border
        for c in range(2, DCOLS + 1):
            ws2.cell(row=current_row, column=c).fill   = _cust_header_fill
            ws2.cell(row=current_row, column=c).border = _border
        ws2.row_dimensions[current_row].height = 22
        current_row += 1

        # ── Transaction rows ─────────────────────────────────────────
        sub_amount = Decimal("0")
        sub_earned = Decimal("0")
        sub_pend   = Decimal("0")
        sub_debt   = Decimal("0")

        for local_idx, tx in enumerate(txs, 1):
            global_idx += 1
            stripe = local_idx % 2 == 0
            fill   = _stripe_fill if stripe else None

            desc = _get_product_summary(tx)

            sub_amount += tx.amount
            sub_earned += tx.bonus_amount
            sub_pend   += tx.pending_bonus
            sub_debt   += tx.debt_change

            _text(ws2, current_row, 1, local_idx,  align=_center, fill=fill)
            _text(ws2, current_row, 2,
                  tx.document_date.strftime("%d.%m.%Y") if tx.document_date else "",
                  align=_center, fill=fill)
            _text(ws2, current_row, 3, tx.document_number, align=_center, fill=fill)
            _text(ws2, current_row, 4, desc, fill=fill)
            _money(ws2, current_row, 5, tx.amount, fill=fill)
            _money(ws2, current_row, 6, tx.bonus_amount if tx.bonus_amount else None,
                   font=_pos_font if tx.bonus_amount > 0 else _data_font, fill=fill)
            _money(ws2, current_row, 7, tx.pending_bonus if tx.pending_bonus else None,
                   font=_warn_font if tx.pending_bonus > 0 else _data_font, fill=fill)
            _money(ws2, current_row, 8, tx.debt_change if tx.debt_change else None,
                   font=_neg_font if tx.debt_change > 0 else _pos_font, fill=fill)
            ws2.row_dimensions[current_row].height = 18
            current_row += 1

        # ── Customer subtotal row ────────────────────────────────────
        ws2.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=3
        )
        sub_lbl = ws2.cell(row=current_row, column=1,
                           value=f"Итого: {cust.full_name or '—'}")
        sub_lbl.font = _summary_font; sub_lbl.fill = _summary_fill
        sub_lbl.alignment = _right;   sub_lbl.border = _border
        for c in range(2, 4):
            ws2.cell(row=current_row, column=c).fill   = _summary_fill
            ws2.cell(row=current_row, column=c).border = _border
        ws2.cell(row=current_row, column=4).fill   = _summary_fill
        ws2.cell(row=current_row, column=4).border = _border
        _money(ws2, current_row, 5, sub_amount, font=_summary_font, fill=_summary_fill)
        _money(ws2, current_row, 6, sub_earned,
               font=Font(name="Calibri", bold=True, size=11, color="375623"), fill=_summary_fill)
        _money(ws2, current_row, 7, sub_pend,
               font=Font(name="Calibri", bold=True, size=11, color="C55A11"), fill=_summary_fill)
        _money(ws2, current_row, 8, sub_debt, font=_summary_font, fill=_summary_fill)
        ws2.row_dimensions[current_row].height = 20
        current_row += 1

        # Blank spacer row
        for c in range(1, DCOLS + 1):
            ws2.cell(row=current_row, column=c).border = _border
        current_row += 1

    # ── Grand total (details sheet) ──────────────────────────────────
    if by_customer:
        ws2.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=4
        )
        gt_lbl = ws2.cell(row=current_row, column=1,
                          value=f"ИТОГО ЗА ПЕРИОД  ({len(by_customer)} клиентов)")
        gt_lbl.font = _total_font; gt_lbl.fill = _total_fill
        gt_lbl.alignment = _right; gt_lbl.border = _border
        for c in range(2, 5):
            ws2.cell(row=current_row, column=c).font   = _total_font
            ws2.cell(row=current_row, column=c).fill   = _total_fill
            ws2.cell(row=current_row, column=c).border = _border
        _money(ws2, current_row, 5, grand_amount, font=_total_font, fill=_total_fill)
        _money(ws2, current_row, 6, grand_earned, font=_total_font, fill=_total_fill)
        _money(ws2, current_row, 7, grand_pend,   font=_total_font, fill=_total_fill)
        ws2.cell(row=current_row, column=8).fill   = _total_fill
        ws2.cell(row=current_row, column=8).border = _border
        ws2.row_dimensions[current_row].height = 24

    _set_col_widths(ws2, {
        "A": 5, "B": 13, "C": 17, "D": 48,
        "E": 18, "F": 18, "G": 18, "H": 18,
    })

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



# ── Styles ────────────────────────────────────────────────────────────────────

_thin = Side(style="thin", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
_title_font = Font(name="Calibri", bold=True, size=14)
_subtitle_font = Font(name="Calibri", size=11, italic=True, color="666666")
_data_font = Font(name="Calibri", size=11)
_money_fmt = '#,##0.00'
_neg_money_font = Font(name="Calibri", size=11, color="CC0000")
_pos_money_font = Font(name="Calibri", size=11, color="228B22")
_center = Alignment(horizontal="center", vertical="center")
_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
_right = Alignment(horizontal="right", vertical="center")
_stripe_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
_summary_font = Font(name="Calibri", bold=True, size=11)
_summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_balance_neg_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
_balance_pos_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _border


def _style_data_cell(cell, is_stripe=False):
    cell.font = _data_font
    cell.border = _border
    if is_stripe:
        cell.fill = _stripe_fill


def _fmt(value):
    """Format number Russian-style: −2 860 333,33"""
    v = Decimal(str(value))
    sign = "−" if v < 0 else ""
    v = abs(v)
    integer_part = int(v)
    decimal_part = f"{v - integer_part:.2f}"[2:]
    s = f"{integer_part:,}".replace(",", " ")
    return f"{sign}{s},{decimal_part}"


# ── Purchase History Excel ────────────────────────────────────────────────────

def generate_purchase_history(customer, transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = "История операций"
    ws.sheet_properties.tabColor = "4472C4"

    # ── Title block ──
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"📋 История операций — {customer.full_name}")
    title_cell.font = _title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    today = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.cell(row=2, column=1, value=f"Сформировано: {today}  |  Телефон: {customer.phone}").font = _subtitle_font
    ws.row_dimensions[2].height = 20

    # ── Balance summary block ──
    balance = customer.debt_balance
    bonus = customer.bonus_balance

    ws.merge_cells("A3:D3")
    bal_fill = _balance_neg_fill if balance < 0 else _balance_pos_fill if balance > 0 else _summary_fill
    bal_cell = ws.cell(row=3, column=1, value=f"💰 Баланс (МойСклад): {_fmt(balance)}")
    bal_cell.font = Font(name="Calibri", bold=True, size=12)
    bal_cell.fill = bal_fill
    bal_cell.border = _border
    for c in range(2, 5):
        ws.cell(row=3, column=c).fill = bal_fill
        ws.cell(row=3, column=c).border = _border

    ws.merge_cells("E3:H3")
    bon_cell = ws.cell(row=3, column=5, value=f"💎 Бонусы: {_fmt(bonus)}")
    bon_cell.font = Font(name="Calibri", bold=True, size=12)
    bon_cell.fill = _summary_fill
    bon_cell.border = _border
    for c in range(6, 9):
        ws.cell(row=3, column=c).fill = _summary_fill
        ws.cell(row=3, column=c).border = _border
    ws.row_dimensions[3].height = 26

    # ── Data table ──
    header_row = 5
    headers = ["№", "Дата", "Тип операции", "Номер документа",
               "Приход (+)", "Расход (−)", "Бонусы", "Описание (товары)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 28

    total_income = Decimal("0")
    total_expense = Decimal("0")
    total_bonus = Decimal("0")

    # Determine which types are "income" vs "expense" from client's perspective
    income_types = {"payment_in", "cash_in", "return"}  # money coming back to client or reducing debt
    expense_types = {"sale", "payment_out", "cash_out"}  # money client pays or owes

    for idx, t in enumerate(transactions, 1):
        row = header_row + idx
        is_stripe = idx % 2 == 0

        is_income = t.type in income_types
        is_expense = t.type in expense_types

        income_val = float(t.amount) if is_income else ""
        expense_val = float(t.amount) if is_expense else ""

        if is_income:
            total_income += t.amount
        elif is_expense:
            total_expense += t.amount

        cells = [
            ws.cell(row=row, column=1, value=idx),
            ws.cell(row=row, column=2, value=t.document_date.strftime("%d.%m.%Y") if t.document_date else ""),
            ws.cell(row=row, column=3, value=t.get_type_display()),
            ws.cell(row=row, column=4, value=t.document_number),
            ws.cell(row=row, column=5, value=income_val),
            ws.cell(row=row, column=6, value=expense_val),
            ws.cell(row=row, column=7, value=float(t.bonus_amount) if t.bonus_amount != 0 else ""),
            ws.cell(row=row, column=8, value=t.description),
        ]

        for c in cells:
            _style_data_cell(c, is_stripe)

        cells[0].alignment = _center
        cells[1].alignment = _center
        cells[2].alignment = _left
        cells[3].alignment = _center

        # Income column — green
        cells[4].alignment = _right
        if isinstance(cells[4].value, (int, float)):
            cells[4].number_format = _money_fmt
            cells[4].font = _pos_money_font

        # Expense column — red
        cells[5].alignment = _right
        if isinstance(cells[5].value, (int, float)):
            cells[5].number_format = _money_fmt
            cells[5].font = _neg_money_font

        # Bonus column
        cells[6].alignment = _right
        if isinstance(cells[6].value, (int, float)):
            cells[6].number_format = _money_fmt
            if cells[6].value > 0:
                cells[6].font = _pos_money_font
            elif cells[6].value < 0:
                cells[6].font = _neg_money_font

        cells[7].alignment = _left

        total_bonus += t.bonus_amount

    # ── Totals row ──
    n = len(transactions)
    tr = header_row + n + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
    lbl = ws.cell(row=tr, column=1, value="ИТОГО")
    lbl.font = _summary_font
    lbl.fill = _summary_fill
    lbl.alignment = _right
    lbl.border = _border
    for c in range(2, 5):
        ws.cell(row=tr, column=c).fill = _summary_fill
        ws.cell(row=tr, column=c).border = _border

    inc_cell = ws.cell(row=tr, column=5, value=float(total_income) if total_income else "")
    inc_cell.font = Font(name="Calibri", bold=True, size=11, color="228B22")
    inc_cell.fill = _summary_fill
    inc_cell.number_format = _money_fmt
    inc_cell.alignment = _right
    inc_cell.border = _border

    exp_cell = ws.cell(row=tr, column=6, value=float(total_expense) if total_expense else "")
    exp_cell.font = Font(name="Calibri", bold=True, size=11, color="CC0000")
    exp_cell.fill = _summary_fill
    exp_cell.number_format = _money_fmt
    exp_cell.alignment = _right
    exp_cell.border = _border

    bon_total = ws.cell(row=tr, column=7, value=float(total_bonus) if total_bonus else "")
    bon_total.font = _summary_font
    bon_total.fill = _summary_fill
    bon_total.number_format = _money_fmt
    bon_total.alignment = _right
    bon_total.border = _border

    ws.cell(row=tr, column=8).fill = _summary_fill
    ws.cell(row=tr, column=8).border = _border

    # ── Column widths ──
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 45
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# (old generate_admin_report removed — replaced by the new version above)